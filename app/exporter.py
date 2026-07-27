from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_excel(rows, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Films"

    headers = [
        "Nom", "Chemin", "Taille (Go)", "Durée",
        "Résolution", "Bitrate vidéo (Mb/s)",
        "Codec vidéo", "Codec audio", "Canaux audio",
        "Langues audio", "Sous-titres", "HDR",
        "Titre IA", "Année IA", "Confiance IA",
        "Statut IA", "Notes IA",
        "Titre TMDb", "Titre original TMDb", "Année TMDb",
        "Score TMDb (%)", "Source comparaison", "Score comparaison (%)", "Statut comparaison", "Nom proposé",
        "Groupe doublon", "Similarité doublon (%)",
        "Score qualité", "Alertes qualité",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in rows:
        resolution = (
            f"{row['width']}x{row['height']}"
            if row["width"] and row["height"] else ""
        )

        duration = ""
        if row["duration"]:
            total = int(row["duration"])
            hours, remainder = divmod(total, 3600)
            minutes, seconds = divmod(remainder, 60)
            duration = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

        ai_confidence = (
            round(float(row["ai_confidence"]) * 100, 1)
            if row["ai_confidence"] is not None else ""
        )
        tmdb_score = (
            round(float(row["tmdb_score"]) * 100, 1)
            if row["tmdb_score"] is not None else ""
        )
        duplicate_score = (
            round(float(row["duplicate_score"]) * 100, 1)
            if row["duplicate_score"] is not None else ""
        )
        bitrate = (
            round(float(row["video_bitrate"]) / 1_000_000, 2)
            if row["video_bitrate"] else ""
        )

        sheet.append([
            row["filename"], row["filepath"],
            round((row["filesize"] or 0) / 1_073_741_824, 2),
            duration, resolution, bitrate,
            row["video_codec"] or "", row["audio_codec"] or "",
            row["audio_channels"] or "", row["audio_languages"] or "",
            row["subtitle_languages"] or "", row["hdr"] or "",
            row["ai_title"] or "", row["ai_year"] or "",
            ai_confidence, row["ai_status"] or "",
            row["ai_notes"] or "", row["tmdb_title"] or "",
            row["tmdb_original_title"] or "", row["tmdb_year"] or "",
            tmdb_score, row["comparison_source"] or "",
            round(float(row["comparison_score"] or 0) * 100, 1) if row["comparison_score"] is not None else "",
            row["comparison_status"] or "", row["proposed_filename"] or "",
            row["duplicate_group"] or "", duplicate_score,
            row["quality_score"] if row["quality_score"] is not None else "",
            row["quality_flags"] or "",
        ])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(filename)
